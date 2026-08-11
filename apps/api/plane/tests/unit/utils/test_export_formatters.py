# Copyright (c) 2023-present Plane Software, Inc. and contributors
# SPDX-License-Identifier: AGPL-3.0-only
# See the LICENSE file for details.

"""Unit tests for the export formatters.

These produce the files users download. Two properties matter beyond "it
serialises": the schema decides column order and headings, so a drifting field
list silently reorders or renames someone's spreadsheet columns; and CSV values
are passed through a formula-injection guard, because a spreadsheet opening an
exported cell starting with = or + will happily execute it.
"""

import csv
import io
import json

import pytest
from openpyxl import load_workbook

from plane.utils.exporters.formatters import (
    BaseFormatter,
    CSVFormatter,
    JSONFormatter,
    XLSXFormatter,
)
from plane.utils.exporters.schemas.base import (
    ExportSchema,
    JSONField,
    ListField,
    StringField,
)


class _RecordSchema(ExportSchema):
    """Field order here is the column order in every export."""

    identifier = StringField(label="ID")
    name = StringField()
    assignee_names = ListField()
    custom_properties = JSONField()


RECORDS = [
    {
        "identifier": "TP-1",
        "name": "First item",
        "assignee_names": ["Ada", "Grace"],
        "custom_properties": {"severity": "high"},
    },
    {
        "identifier": "TP-2",
        "name": "Second item",
        "assignee_names": [],
        "custom_properties": None,
    },
]


@pytest.mark.unit
class TestBaseFormatter:
    def test_format_is_abstract(self):
        with pytest.raises(NotImplementedError):
            BaseFormatter().format("f", [], _RecordSchema)

    def test_field_info_follows_declaration_order(self):
        order, labels = BaseFormatter._get_field_info(_RecordSchema)
        assert order == ["identifier", "name", "assignee_names", "custom_properties"]

    def test_explicit_label_wins(self):
        _, labels = BaseFormatter._get_field_info(_RecordSchema)
        assert labels["identifier"] == "ID"

    def test_labels_fall_back_to_a_title_cased_field_name(self):
        _, labels = BaseFormatter._get_field_info(_RecordSchema)
        assert labels["assignee_names"] == "Assignee Names"

    def test_a_schema_without_declared_fields_is_rejected(self):
        class NotASchema:
            pass

        with pytest.raises(ValueError, match="_declared_fields"):
            BaseFormatter._get_field_info(NotASchema)


@pytest.mark.unit
class TestCSVFormatter:
    def _rows(self, content):
        return list(csv.reader(io.StringIO(content)))

    def test_extension_is_appended_to_the_filename(self):
        name, _ = CSVFormatter().format("export", RECORDS, _RecordSchema)
        assert name == "export.csv"

    def test_no_records_produces_an_empty_body(self):
        assert CSVFormatter().format("export", [], _RecordSchema) == ("export.csv", "")

    def test_header_row_uses_the_schema_labels(self):
        _, content = CSVFormatter().format("export", RECORDS, _RecordSchema)
        assert self._rows(content)[0] == ["ID", "Name", "Assignee Names", "Custom Properties"]

    def test_lists_are_joined_and_dicts_serialised(self):
        _, content = CSVFormatter().format("export", RECORDS, _RecordSchema)
        first = self._rows(content)[1]
        assert first[2] == "Ada, Grace"
        assert json.loads(first[3]) == {"severity": "high"}

    def test_list_joiner_is_configurable(self):
        _, content = CSVFormatter().format(
            "export", RECORDS, _RecordSchema, options={"list_joiner": " | "}
        )
        assert self._rows(content)[1][2] == "Ada | Grace"

    def test_none_becomes_an_empty_cell(self):
        _, content = CSVFormatter().format("export", RECORDS, _RecordSchema)
        assert self._rows(content)[2][3] == ""

    def test_requested_fields_narrow_the_columns(self):
        _, content = CSVFormatter().format(
            "export", RECORDS, _RecordSchema, options={"fields": ["identifier", "name"]}
        )
        assert self._rows(content)[0] == ["ID", "Name"]

    def test_formula_injection_is_neutralised(self):
        """A cell starting with a formula character must not stay executable."""
        record = [{"identifier": "=1+1", "name": "x", "assignee_names": [], "custom_properties": {}}]
        _, content = CSVFormatter().format("export", record, _RecordSchema)
        assert not self._rows(content)[1][0].startswith("=")


@pytest.mark.unit
class TestJSONFormatter:
    def test_no_records_produces_an_empty_array(self):
        assert JSONFormatter().format("export", [], _RecordSchema) == ("export.json", "[]")

    def test_extension_and_labelled_keys(self):
        name, content = JSONFormatter().format("export", RECORDS, _RecordSchema)
        assert name == "export.json"
        assert json.loads(content)[0]["ID"] == "TP-1"

    def test_types_are_preserved_rather_than_stringified(self):
        """Unlike the flat formats, JSON keeps lists and objects intact."""
        _, content = JSONFormatter().format("export", RECORDS, _RecordSchema)
        first = json.loads(content)[0]
        assert first["Assignee Names"] == ["Ada", "Grace"]
        assert first["Custom Properties"] == {"severity": "high"}

    def test_absent_keys_are_omitted_rather_than_nulled(self):
        _, content = JSONFormatter().format("export", [{"identifier": "TP-9"}], _RecordSchema)
        assert json.loads(content) == [{"ID": "TP-9"}]

    def test_requested_fields_narrow_the_keys(self):
        _, content = JSONFormatter().format(
            "export", RECORDS, _RecordSchema, options={"fields": ["identifier"]}
        )
        assert list(json.loads(content)[0]) == ["ID"]


@pytest.mark.unit
class TestXLSXFormatter:
    def _sheet(self, content):
        return load_workbook(io.BytesIO(content)).active

    def test_no_records_still_produces_a_readable_workbook(self):
        name, content = XLSXFormatter().format("export", [], _RecordSchema)
        assert name == "export.xlsx"
        assert self._sheet(content).max_row == 1  # openpyxl reports one empty row

    def test_header_and_values(self):
        _, content = XLSXFormatter().format("export", RECORDS, _RecordSchema)
        sheet = self._sheet(content)
        assert [c.value for c in sheet[1]] == [
            "ID",
            "Name",
            "Assignee Names",
            "Custom Properties",
        ]
        assert sheet["A2"].value == "TP-1"

    def test_lists_and_dicts_are_flattened_to_text(self):
        _, content = XLSXFormatter().format("export", RECORDS, _RecordSchema)
        sheet = self._sheet(content)
        assert sheet["C2"].value == "Ada, Grace"
        assert json.loads(sheet["D2"].value) == {"severity": "high"}

    def test_requested_fields_narrow_the_columns(self):
        _, content = XLSXFormatter().format(
            "export", RECORDS, _RecordSchema, options={"fields": ["identifier"]}
        )
        assert self._sheet(content).max_column == 1
